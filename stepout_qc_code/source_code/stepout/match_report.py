import os
import time
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

src_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
parent_dir = os.path.abspath(os.path.join(src_dir, os.pardir))
os.chdir(parent_dir)


class MatchReport:

    path = r'match_report'
    if not os.path.exists(path):
        os.makedirs(path)

    def __init__(self, data, df):
        self.df = df
        self.data = data
        self.match_id = self.data['match_id']
        self.team_a_name = self.data['team_a_name']
        self.team_b_name = self.data['team_b_name']
        self.match_report_path_xlsx = "match_report/Match_ID_" + str(self.match_id) + "_Time_" \
                                 + time.strftime("%H-%M-%S") + ".xlsx"

    def write_match_report(self, final_shot_pass_df, sheet_name):
        # writing the match report into an excel file
        if os.path.exists(self.match_report_path_xlsx):
            wb = load_workbook(self.match_report_path_xlsx)
            ws = wb.create_sheet(sheet_name)

        else:
            wb = Workbook()
            ws = wb.active
            ws.title = f'{sheet_name}'

        for r in dataframe_to_rows(final_shot_pass_df, index=False, header=True):
            ws.append(r)

        for cell in ws[1]:
            cell.style = 'Pandas'

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 25

        assist_goal_list = ['Through Pass Assist', 'Short Pass Assist', 'Long Pass Assist', 'Cross Assist',
                            'Close Shot Goal', 'Long Shot Goal', 'Header Goal']

        # coloring some cells based on a condition
        for row_idx in range(2, final_shot_pass_df.shape[0] + 2):
            action_taken_cell = ws.cell(row=row_idx, column=4)
            if action_taken_cell.value in assist_goal_list:
                action_taken_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        wb.save(self.match_report_path_xlsx)
        return True

    def reorder_col(self, shot_pass_df):
        # dropping and reordering columns
        shot_pass_df.drop(columns=['action', 'notation', 'start_grid', 'end_grid', 'foot', 'team'], inplace=True)

        desired_order = ['timestamp', 'jersey_number', 'special_attribute', 'action_taken', 'receiver', 'half']

        shot_pass_df = shot_pass_df[desired_order]

        return shot_pass_df

    def special_attribute(self, shot_pass_df):
        # creating special_attribute col
        sp_att_dict = {
            'X': np.nan,
            'FK': 'Free Kick',
            'PK': 'Penalty Kick',
            'GK': 'Goal Kick',
            'CN': 'Corner Kick'
        }
        shot_pass_df['special_attribute'] = shot_pass_df['special_attribute'].map(sp_att_dict)
        return self.reorder_col(shot_pass_df)

    def half_col(self, shot_pass_df):
        # creating 'half' col
        half_dict = {
            'FHN': 'First Half',
            'FHI': 'First Half Injury Time',
            'SHN': 'Second Half',
            'SHI': 'Second Half Injury Time',
            'ET1N': 'Extra-Time First Half',
            'ET1I': 'Extra-Time First Half Injury Time',
            'ET2N': 'Extra-Time Second Half',
            'ET2I': 'Extra-Time Second Half Injury Time',
            'PK': 'Penalty Shoot-Out'
        }
        shot_pass_df['half'] = shot_pass_df['half'].map(half_dict)
        return self.special_attribute(shot_pass_df)

    def action_taken_col(self, shot_pass_df):
        # creating 'action_taken' col
        action_dict = {
            'TB': {'2': 'Key Through Pass', '3': 'Through Pass Assist'},
            'SP': {'2': 'Key Short Pass', '3': 'Short Pass Assist'},
            'LP': {'2': 'Key Long Pass', '3': 'Long Pass Assist'},
            'C': {'2': 'Key Cross', '3': 'Cross Assist'},
            'CS': {'0': 'Off-Target Close Shot',
                   '1': 'Simple Close Shot',
                   '2': 'CS-Hitting Cross bar/Post',
                   '3': 'Brilliant Close Shot',
                   '4': 'Close Shot Goal'},
            'LS': {'0': 'Off-Target Long Shot',
                   '1': 'Simple Long Shot',
                   '2': 'LS-Hitting Cross bar/Post',
                   '3': 'Brilliant Long Shot',
                   '4': 'Long Shot Goal'},
            'H': {'0': 'Off-Target Header',
                  '1': 'Simple Header',
                  '2': 'Header-Hitting Cross bar/Post',
                  '3': 'Brilliant Header',
                  '4': 'Header Goal'}
        }

        def action_to_words(row):
            return action_dict[row['action']][row['notation']]

        shot_pass_df['action_taken'] = shot_pass_df[['action', 'notation']].apply(action_to_words, axis=1)

        return self.half_col(shot_pass_df)


    def receiver_col(self, shot_pass_df):
        """
        creating 'receiver' col
        :param shot_pass_df: pd.DataFrame
        :return: None
        """
        index_list = []

        for index, row in shot_pass_df.iterrows():
            if row['action'] in ['XSP', 'XLP', 'XC', 'XTB']:
                shot_pass_df.at[index - 1, 'receiver'] = row['jersey_number']
                index_list.append(index)

        shot_pass_df.drop(index_list, inplace=True)

        return self.action_taken_col(shot_pass_df)


    def key_assist_shot_detail(self, team, team_name):
        """
        Creating a df made key pass, assist, shots and goals
        :param team: str -> 'A' or 'B'
        :param team_name: str -> team name
        :return: None
        """
        key_assist_pass_mask = ((self.df['team'] == team) & \
                                (self.df['action'].isin(['SP', 'LP', 'C', 'TB', 'XSP', 'XLP', 'XC', 'XTB']))& \
                                (self.df['notation'].isin(['2', '3'])))
        shot_goal_mask = ((self.df['team'] == team) & (self.df['action'].isin(['CS', 'LS', 'H'])))
        shot_pass_df = self.df[key_assist_pass_mask | shot_goal_mask].copy()

        final_shot_pass_df = self.receiver_col(shot_pass_df)
        sheet_name = team_name

        bool = self.write_match_report(final_shot_pass_df, sheet_name)

        if bool:
            print(f"{team_name} match report created")

    def calling_func(self):
        self.key_assist_shot_detail('A', self.team_a_name)
        self.key_assist_shot_detail('B', self.team_b_name)













