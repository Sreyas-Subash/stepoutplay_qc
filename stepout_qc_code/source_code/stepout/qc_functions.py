import os
import time
from options import option

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils.cell import coordinate_from_string
from package_necessity import star_count
from openpyxl.styles import PatternFill

src_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
parent_dir = os.path.abspath(os.path.join(src_dir, os.pardir))
os.chdir(parent_dir)

class QualityChecks:
    flag = True
    path = r'excel_logs'
    if not os.path.exists(path):
        os.makedirs(path)

    def __init__(self, df, match_id, qc_path_xlsx):
        self.df = df
        self.match_id = match_id
        # self.qc_path_xlsx = "excel_logs/Match_ID_" + str(match_id) + "_Time_" + time.strftime("%H-%M-%S") + ".xlsx"
        self.qc_path_xlsx = qc_path_xlsx
        self.gd_ad_idx_list = []

    def qc_excel_log(self, df, message):
        """
        Records all the errors to an excel file
        :param df: pd.DataFrame
        :param message: str
        :param size: int
        :param sheet_name: str
        :return: None
        """

        # Create a PatternFill object for a solid red fill
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        if not os.path.exists(self.qc_path_xlsx):
            wb = Workbook()
            ws = wb.active
            ws.title = 'All_QC'
            match_id_cell = ws.cell(row=1, column=1, value=f'match_id = {self.match_id}')
            match_id_cell.fill = red_fill
            self.flag = False

        else:
            wb = load_workbook(self.qc_path_xlsx)
            ws = wb['All_QC']
            if self.flag:
                match_id_cell = ws.cell(row=ws.max_row+2, column=1, value=f'match_id = {self.match_id}')
                match_id_cell.fill = red_fill
                m_id = self.match_id
                self.flag = False

        # Find the max row in the existing sheet to append data
        current_max_row = ws.max_row
        # Write the message in the first cell
        msg_cell = ws.cell(row=current_max_row+1, column=1, value=f'{message}')
        msg_cell.font = Font(size=12, underline='single')
        msg_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Set the width of the first column
        ws.column_dimensions['A'].width = 45

        if df.shape[0] != 0:
            # Write the DataFrame to the sheet starting from the next empty row
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)

            # Apply styles to the new data
            for cell in ws[current_max_row + 2]:
                cell.style = 'Pandas'

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        max_cell = cell

            coordinate = coordinate_from_string(max_cell.coordinate)
            max_col = coordinate[0]
            max_row = coordinate[1]

            ws.move_range(f"A{current_max_row+2}:{max_col}{max_row}", rows=0, cols=1)

        # Save the workbook
        wb.save(self.qc_path_xlsx)

        # if os.path.exists(self.qc_path_xlsx):
        #     wb = load_workbook(self.qc_path_xlsx)
        #     ws = wb.create_sheet(f'{sheet_name}')
        #
        # else:
        #     wb = Workbook()
        #     ws = wb.active
        #     ws.title = f'{sheet_name}'
        #
        # a1 = ws['A1']
        # a1.value = f'{message}'
        #
        # a1.font = Font(size=12, underline='single')
        # a1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        #
        # if df.shape[0] != 0:
        #     for row in dataframe_to_rows(df, index=False, header=True):
        #         ws.append(row)
        #
        #     for cell in ws[2]:
        #         cell.style = 'Pandas'
        #
        #     for row in ws.iter_rows():
        #         for cell in row:
        #             if cell.value:
        #                 max_cell = cell
        #
        #     coordinate = coordinate_from_string(max_cell.coordinate)
        #     max_col = coordinate[0]
        #     max_row = coordinate[1]
        #
        #     ws.move_range(f"A2:{max_col}{max_row}", rows=0, cols=1)
        #
        # ws.column_dimensions['A'].width = size
        #
        # wb.save(self.qc_path_xlsx)

    def display_output(self, output_df, message, sheet_name, wrong_data_count = 1):
        if option.lower() == 'm':
            if wrong_data_count != 0:
                self.qc_excel_log(output_df, message)
        else:
            if wrong_data_count != 0:
                print('*' * star_count)
                print(message)
                print(output_df.to_string())
                print('*' * star_count)
                self.qc_excel_log(output_df, message)
            else:
                print(f'###{sheet_name} qc done###')

    def non_def_foul(self):
        """
        This function checks if -
        A non-defensive action(other than ST, SL, AD, GD, HB, THW) was given a foul
        :return: None
        """
        mask = ~(self.df['action'].isin(['ST', 'SL', 'AD', 'GD', 'HB', 'THW'])) & (self.df['special_attribute'].isin(['F', 'YC', 'RC']))
        non_df_action = self.df[mask]
        non_df_action_count = non_df_action['action'].count()
        message = 'A non defensive action was tagged as a foul. Please check!'
        self.display_output(non_df_action, message, 'non_df_action_foul', non_df_action_count)

    def successful_def(self):
        """
         This function checks if -
         A successful defensive action was tagged as a foul
        :return: None
        """
        mask = (self.df['action'].isin(['ST', 'SL', 'AD', 'GD', 'THW']))  & (self.df['notation'] == '1') & \
               (self.df['special_attribute'].isin(['F', 'YC', 'RC']))
        successful_df_action = self.df[mask]
        successful_df_action_foul_count = successful_df_action['action'].count()
        message = 'A successful defensive action was tagged as a foul. Please check!'
        self.display_output(successful_df_action, message, 'successful_df_action_foul', successful_df_action_foul_count)

    def corner_qc_1(self):
        """
        The purpose of this function -
        If there are any corners which didn't start from the correct corner start grid(1,01,71)
        """

        mask = (self.df['special_attribute'] == 'CN') & ~(self.df['start_grid'].isin(['01', '1', '71']))
        false_cn = self.df[mask]
        false_cn_count = false_cn['action'].count()
        message = 'This Corner started from a false start grid(not starting from [1, 01, 71]). Please check!'
        self.display_output(false_cn, message, 'corner_grid_check', false_cn_count)

    def corner_qc_2(self):
        """
        The purpose of this function -
        If there are any unassigned corners within the crosses
        """
        # To check if any crosses starting from (1, 01, 71) are corners
        mask = (self.df['special_attribute'] != 'CN') \
               & (self.df['start_grid'].isin(['1','01','71'])) \
               & (self.df['action'] == 'C')
        corner_in_cross = self.df[mask]
        corner_in_cross_count = corner_in_cross['action'].count()
        message = f'{corner_in_cross_count} cross found with corner grid. Check if they are corners'
        self.display_output(corner_in_cross, message, 'cross_check', corner_in_cross_count)

    def gk_qc_1(self):
        """
        This function checks if GK is taken from outside goal area
        :return: None
        """

        mask = (self.df['special_attribute'] == 'GK') & ~(self.df['start_grid'].isin(['40', '50']))
        wrong_goalkick = self.df[mask]
        wrong_goalkick_count = wrong_goalkick['action'].count()
        message = 'GK QC-1\nThis goal kick is taken outside the goal area. Please check!'
        self.display_output(wrong_goalkick, message, 'GK QC-1', wrong_goalkick_count)

    def gk_qc_2(self):
        """
        This function checks if GH or GT is taken from outside penalty area
        :return:
        """

        gk_d_grids = ['29', '30', '39', '40', '49', '50', '59', '60']
        mask = (self.df['action'].isin(['GH', 'GT'])) & ~(self.df['start_grid'].isin(gk_d_grids))
        wrong_goalkeeper = self.df[mask]
        wrong_goalkeeper_count = wrong_goalkeeper['action'].count()
        message = 'GK QC-2\nThis GH or GT taken outside penalty area. Please check!'
        self.display_output(wrong_goalkeeper, message, 'GK QC-2', wrong_goalkeeper_count)

    def penalty_qc(self):
        """
        This function checks if PK is taken from the appropriate grid location(32, 42)
        :return: None
        """
        mask = (self.df['special_attribute'] == 'PK') & ~(self.df['start_grid'].isin(['32', '42']))
        wrong_pk = self.df[mask]
        wrong_pk_count = wrong_pk['action'].count()
        message = 'This penalty kick is taken outside the penalty area(32, 42). Please check!'
        self.display_output(wrong_pk, message, 'Penalty_check', wrong_pk_count)

    def unsuccessful_interception(self):
        """
        This function checks if
        There are any unsuccessful interceptions(IN-0).
        :return: None
        """
        mask = (self.df['action'].isin(['IN', 'XIN'])) & (self.df['notation'] == '0')
        unsuccessful_interception = self.df[mask]
        unsuccessful_interception_count = unsuccessful_interception['action'].count()
        message = 'An unsuccessful interception was tagged. Please check!'
        self.display_output(unsuccessful_interception, message, 'unsuccessful_interception_check', unsuccessful_interception_count)

    def misbehaviour_foul_count(self):
        """
        This function finds the misbehaviour foul counts + THW-0-F(Foul throws)
        :return:
        misbehaviour_foul_a : int -> misbehaviour foul count for team A
        misbehaviour_foul_b : int -> misbehaviour foul count for team B
        """
        self.misbehaviour_foul_index = []
        misbehaviour_foul_a = 0
        misbehaviour_foul_b = 0
        # misbehaviour foul standard notation ST-0 with YC or RC
        mask_1 = (self.df['action'] == 'ST') & (self.df['notation'] == '0') & (self.df['special_attribute'].isin(['YC', 'RC']))
        mask_2 = (self.df['action'] == 'THW') & (self.df['notation'] == '0') & (self.df['special_attribute'].isin(['F']))
        misbehaviour_foul_index_list = self.df[mask_1 | mask_2].index

        for index in misbehaviour_foul_index_list:
            # misbehaviour foul standard notation will not have XST
            # df.index[-1] to make sure index does not run out of range
            mask_3 = (index == self.df.index[-1]) or ((self.df.loc[index + 1, 'action'] != 'XST') & (self.df.loc[index + 1, 'action'] != 'XTHW'))
            if (mask_3) & (self.df.loc[index, 'team'] == 'A'):
                misbehaviour_foul_a += 1
                self.misbehaviour_foul_index.append(index)
            elif (mask_3) & (self.df.loc[index, 'team'] == 'B'):
                misbehaviour_foul_b += 1
                self.misbehaviour_foul_index.append(index)

        return misbehaviour_foul_a, misbehaviour_foul_b

    def fk_pk_foul_count(self):
        """
        finding fk-pk and fouls count
        :return:
        teama_foul : int -> team A foul total count after misbehaviour foul count is subtracted
        teama_fk_pk : int -> team A fk-pk total count
        teamb_foul : int -> team B foul total count after misbehaviour foul count is subtracted
        teamb_fk_pk : int -> team B fk-pk total count
        """
        misbehaviour_foul_a, misbehaviour_foul_b = self.misbehaviour_foul_count()
        foul_type = (self.df['action'].isin(['HB', 'OFF'])) | (self.df['special_attribute'].isin(['F', 'YC', 'RC']))
        fk_pk = self.df['special_attribute'].isin(['FK', 'PK'])

        teamb_foul = self.df[(self.df['team'] == 'B') & (foul_type)]['action'].count()
        teamb_foul = teamb_foul - misbehaviour_foul_b
        teamb_fk_pk = self.df[(self.df['team'] == 'B') & (fk_pk)]['action'].count()

        teama_foul = self.df[(self.df['team'] == 'A') & (foul_type)]['action'].count()
        teama_foul = teama_foul - misbehaviour_foul_a
        teama_fk_pk = self.df[(self.df['team'] == 'A') & (fk_pk)]['action'].count()

        return teama_foul, teama_fk_pk, teamb_foul, teamb_fk_pk

    def fk_pk_foul_output_df(self, foul_team, fk_pk_team):
        """
        finding the fk-pk or foul without their respective pairs
        :param foul_team: List[str] -> team(s) with wrong fouls
        :param fk_pk_team: List[str] -> team(s) with wrong fk-pk
        :return: self.df.iloc[error_list]: pd.DataFrame -> output dataframe with errors
        """
        foul_mask = ((self.df['team'].isin(foul_team)) & \
                     ((self.df['action'].isin(['HB', 'OFF'])) | \
                      (self.df['special_attribute'].isin(['F', 'YC', 'RC']))))
        fk_mask = ((self.df['team'].isin(fk_pk_team)) & (self.df['special_attribute'].isin(['FK', 'PK'])))
        fk_pk_foul_df = self.df[foul_mask | fk_mask].copy()

        foul_index_list = list(self.df[foul_mask].index)
        foul_index_list = [x for x in foul_index_list if x not in self.misbehaviour_foul_index]

        fk_pk_index_list = list(self.df[fk_mask].index)
        fk_pk_foul_df_index_list = list(fk_pk_foul_df.index)
        correct_pair_index_list = []

        # checking if an index in fk_pk_foul_df is in foul_index and its next index is in fk_pk_index
        # also checking fk-pk and foul for alternate teams
        for idx in fk_pk_foul_df_index_list[:-1]:
            if idx in foul_index_list:
                foul_index = fk_pk_foul_df_index_list.index(idx)
                fk_pk_index_val = fk_pk_foul_df_index_list[foul_index + 1]
                foul_teamm = self.df.at[idx, 'team']
                if (fk_pk_index_val in fk_pk_index_list) and (self.df.at[fk_pk_index_val, 'team'] != foul_teamm):
                    correct_pair_index_list.append(idx)
                    correct_pair_index_list.append(fk_pk_index_val)

        # removing pairs to get fk-pk or fouls without their pairs
        for val in correct_pair_index_list:
            fk_pk_foul_df_index_list.remove(val)

        # removing misbehaviour fouls
        fk_pk_foul_df_index_list = [x for x in fk_pk_foul_df_index_list if x not in self.misbehaviour_foul_index]
        error_list = fk_pk_foul_df_index_list

        return self.df.iloc[error_list]

    def fk_pk_foul_check(self):
        """
        This function finds the misbehaviour foul counts, subtracts them from total fouls and checks if total fouls equal
        total freekick-penalty
        :return: None
        """
        teama_foul, teama_fk_pk, teamb_foul, teamb_fk_pk = self.fk_pk_foul_count()

        # verifying if the foul and fk-pk count match
        if (teamb_foul == teama_fk_pk) & (teama_foul == teamb_fk_pk):
            # print('###foul_fk-pk qc done###')
            pass
        else:
            teamB = (teamb_foul == teama_fk_pk)
            teamA = (teama_foul == teamb_fk_pk)
            if not (teamB | teamA):
                foul_team = ['A', 'B']
                fk_pk_team = ['A', 'B']
            elif teamB:
                foul_team = ['A']
                fk_pk_team = ['B']
            else:
                foul_team = ['B']
                fk_pk_team = ['A']

            output_df = self.fk_pk_foul_output_df(foul_team, fk_pk_team)
            message = 'Foul to fk-pk not equal\n'\
                    f'Team A FK-PK = {teama_fk_pk}, Team B Fouls = {teamb_foul}\n'\
                    f'Team B FK-PK = {teamb_fk_pk}, Team A Fouls = {teama_foul}'
            self.display_output(output_df, message, 'fk_pk_foul_check')

    def receiver_not_same(self):
        """
        This function checks if the current and receiver action are done by the same player
        :return: None
        """
        error_df = pd.DataFrame()
        receiver_actions = ['ST','SL','IN','GD','AD','SP','LP','TB','C','DR','GT','THW']
        for action in receiver_actions:
            mask = self.df['action'].isin([action, f'X{action}'])
            new_df = self.df[mask]

            for tmstmp in new_df['timestamp'].unique():
                if new_df[new_df['timestamp'] == tmstmp]['team'].count() == 1:
                    continue
                elif len(new_df[new_df['timestamp'] == tmstmp]['team'].unique()) == 1 and \
                        len(new_df[new_df['timestamp'] == tmstmp]['jersey_number'].unique()) == 1:
                    error_df = pd.concat([error_df, new_df[new_df['timestamp'] == tmstmp]])

        message = 'These current and receiver actions are done by the same player. Check!'
        self.display_output(error_df, message, 'current_receiver_same', error_df.shape[0])

    def current_next_action_not_same(self):
        """
        This function checks if the current action and next action are done by the same player
        Only for SP, LP, TB, C
        :return:
        """
        error_df = pd.DataFrame()
        receiver_actions = ['XSP','XLP','XTB','XC']
        mask = self.df['action'].isin(receiver_actions) & self.df['notation'] != 0
        receiver_actions_index_list = self.df[mask].index.tolist()
        for idx in receiver_actions_index_list:
            if idx == self.df.index[-1]:
                continue
            elif (self.df.at[idx - 1,'team'] == self.df.at[idx  +1,'team']) and \
                    (self.df.at[idx - 1, 'jersey_number'] == self.df.at[idx + 1, 'jersey_number']):
                error_df = pd.concat([error_df, self.df.loc[[idx - 1, idx, idx + 1]]])

        message = 'These current and next actions are done by the same player(3 rows are 1 error). Check!'
        self.display_output(error_df, message, 'current_next_same', error_df.shape[0])


    def fhn_shn_absent(self):
        half_list = self.df['half'].unique()
        error_count = 0
        message = ''
        if ('FHN' not in half_list):
            message = "Match completely tagged with SHN"
            error_count = 1
        elif ('SHN' not in half_list):
            message = "Match completely tagged with FHN"
            error_count = 1
        error_df = pd.DataFrame()
        self.display_output(error_df, message, 'fhn_or_shn_absent', error_count)

    def num_of_actions(self):
        actions = self.df.shape[0]
        message = f"Number of actions = {actions}"
        error_df = pd.DataFrame()
        self.display_output(error_df, message, 'action_numbers')

    def gd_ad_idx_list_func(self, gd_ad_df):
        # finding unsuccessful AD and GD, then appending the index of last 2 rows to be corrected
        if '1' not in gd_ad_df['notation'].unique():
            for i in gd_ad_df.iloc[[-2,-1]].index:
                self.gd_ad_idx_list.append(i)

    def gd_ad_correction(self):
        """
        This function writes GD-0 and XGD-0 into GD-1 and XGD-1(likewise for AD) when an unsuccessful GD or AD
        takes place
        :param df: pd.DataFrame -> df to be corrected
        :return: None
        """
        for filter in [['GD', 'XGD'], ['AD', 'XAD']]:
            # finding unique timestamp of all AD and GD
            tmsp_list = self.df[(self.df['action'].isin(filter))]['timestamp'].unique()

            for tmsp in tmsp_list:
                # creating a 4 row df for each GD and AD
                gd_ad_df = self.df[(self.df['action'].isin(filter)) & (self.df['timestamp'] == tmsp)]
                row_count = gd_ad_df['action'].count()
                # if more than one GD or AD takes place at the same timestamp, it becomes a df of more than 4 rows
                if row_count > 4:
                    # making them separate dfs of 4 rows
                    divs = int(row_count / 4)
                    i = 1
                    while i <= divs:
                        gd_ad_dff = gd_ad_df.iloc[4 * (i - 1): 4 * i]
                        i += 1
                        self.gd_ad_idx_list_func(gd_ad_dff)
                else:
                    self.gd_ad_idx_list_func(gd_ad_df)

        error_df = self.df.iloc[self.gd_ad_idx_list]
        message = 'The receiver actions of these unsuccessful AD or GD are not correct(2 rows -> 1 error). Check!'
        self.display_output(error_df, message, 'GD_AD_wrong', error_df.shape[0])



    def calling_func(self):
        self.non_def_foul()
        self.successful_def()
        self.corner_qc_1()
        self.corner_qc_2()
        self.gk_qc_1()
        self.gk_qc_2()
        self.penalty_qc()
        self.unsuccessful_interception()
        self.fk_pk_foul_check()
        self.receiver_not_same()
        # self.current_next_action_not_same()
        self.gd_ad_correction()
        self.fhn_shn_absent()
        # self.num_of_actions()


# if __name__ == "__main__":
#     read_obj = Read()
#     df = read_obj.read()
#
#     match_obj = MatchDetails
#
#     qc_obj = QualityChecks(df)
#     qc_obj.calling_func()



