# pip install openpyxl

import os
import datetime
import pandas as pd
from analyst_details import AnalystDetails
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

src_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
parent_dir = os.path.abspath(os.path.join(src_dir, os.pardir))
os.chdir(parent_dir)


class MatchDetails:

    def __init__(self, analyst_names, match_id):
        self.analyst_names = analyst_names
        self.match_id = match_id
        self.analyst_id = int(input(f'Choose the Analyst ID\n{self.analyst_names} = '))
        self.analyst_name = self.analyst_names[self.analyst_id]
        self.team_a_name = input('Team A name = ')
        self.team_b_name = input('Team B name = ')
        self.current_date = datetime.date.today()

    def game_time_pay(self):
        """
        Assigns game time and remuneration
        :return: None
        """
        self.game_time = int(input("Game time?\n"
                                    "Type 1 for 90 minutes\n"
                                    "Type 2 for 45 minutes\n"
                                    "Type 3 for less than 60 minutes\n"
                                    "Type 4 for manually entering\n"
                                    "--->"))

        if self.game_time == 1:
            self.renumeration = 500
            self.game_time = 90
        elif self.game_time == 2:
            self.renumeration = 250
            self.game_time = 45
        elif self.game_time == 3:
            self.renumeration = 300
            self.game_time = 'Less than 60'
        elif self.game_time == 4:
            self.renumeration = int(input("Enter renumeration : "))
            self.game_time = int(input("Enter game time : "))
        else:
            print('You have chosen an invalid option. Try again!')

        return self.renumeration, self.game_time

    def match_renum_excel(self, write_path_xlsx, data):
        """
        Creates a excel file with match details or adds match details to the existing excel file
        :param write_path_xlsx: str
        :param data: pd.DataFrame
        :return: None
        """
        try:
            self.write_path_xlsx = write_path_xlsx
            self.data = data

            # There are 4 possible scenarios here
            # 1. The file exists
            if os.path.exists(self.write_path_xlsx):
                wb = load_workbook(self.write_path_xlsx)
                sheet_name = self.analyst_name
                sheet_exists = sheet_name in wb.sheetnames

                # 2. File exists but sheet doesn't
                if not sheet_exists:
                    ws = wb.create_sheet(sheet_name)
                    renumeration_df = pd.DataFrame(self.data, index=[0])
                    for r in dataframe_to_rows(renumeration_df, index=False, header=True):
                        ws.append(r)
                    for cell in ws[1]:
                        cell.style = 'Pandas'
                    wb.save(self.write_path_xlsx)

                # 3. File and sheet exists
                else:
                    renumeration_df = pd.read_excel(self.write_path_xlsx, sheet_name=self.analyst_name)
                    renumeration_df.loc[len(renumeration_df)] = self.data
                    if renumeration_df['match_id'].duplicated().sum() > 0:
                        print(renumeration_df[renumeration_df['match_id'].duplicated()])
                        print('This match id seems to be a duplicate. Please check!')

                    else:
                        renumeration_df = pd.DataFrame(self.data, index=[0])
                        for r in dataframe_to_rows(renumeration_df, index=False, header=False):
                            ws = wb[sheet_name]
                            ws.append(r)
                        wb.save(self.write_path_xlsx)

            # 4. File does not exists
            else:
                renumeration_df = pd.DataFrame(self.data, index=[0])
                renumeration_df.to_excel(self.write_path_xlsx, sheet_name=self.analyst_name, index=False)

        except PermissionError as e:
            print(f'{e}\nThe match_renum_details xlsx file might be open. Please close it and try again')

        except Exception as e:
            print(f'An error occured : {e}')

    def calling_func(self):
        self.renumeration, self.game_time = self.game_time_pay()

        self.data = {'team_a_name': self.team_a_name,
                     'team_b_name': self.team_b_name,
                     'match_id': self.match_id,
                     'game_time': self.game_time,
                     'current_date': self.current_date,
                     'renumeration': self.renumeration
                     }

        write_path_xlsx = r'write_string\match_renum_detials.xlsx'

        self.match_renum_excel(write_path_xlsx, self.data)

        print("*"*50)
        print("Match details added")
        print("*"*50)


if __name__ == "__main__":
    analyst_obj = AnalystDetails()
    analyst_names = analyst_obj.analyst_names


    match_obj = MatchDetails(analyst_names)
    match_obj.calling_func()